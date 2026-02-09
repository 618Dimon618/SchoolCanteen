import os
from datetime import date, datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import db, User, MenuItem, Category, Order, OrderItem, Notification, PurchaseRequest, MenuItemIngredient, Product, Allergy, MenuItemAllergy, Subscription, Payment, Review, Favorite
from db_functions import *

app = Flask(__name__)
app.secret_key = 'school_canteen_secret_key_2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///canteen.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)


@app.after_request
def add_no_cache_headers(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, post-check=0, pre-check=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


ROLE_REQUIRED = {
    'student': 'student',
    'order': 'student',
    'buy_subscription': 'student',
    'add_balance_route': 'student',
    'receive_order': 'student',
    'profile': None,
    'toggle_allergy': None,
    'notifications': None,
    'reviews': None,
    'add_review_route': None,
    'toggle_favorite': None,
    'cook': 'cook',
    'cook_dishes': 'cook',
    'toggle_item': 'cook',
    'create_dish': 'cook',
    'delete_dish': 'cook',
    'cook_item': 'cook',
    'prepare_order': 'cook',
    'add_request': 'cook',
    'cook_issued': 'cook',
    'admin': 'admin',
    'approve': 'admin',
    'reject': 'admin',
    'approve_user': 'admin',
    'reject_user': 'admin',
    'download_report': 'admin',
}

ROLE_HOME = {
    'student': 'student',
    'cook': 'cook',
    'admin': 'admin',
}


@app.before_request
def check_session():
    public_routes = ['login', 'register', 'static', 'index']
    endpoint = request.endpoint

    if not endpoint:
        return

    if endpoint in public_routes:
        return

    if 'user_id' not in session:
        flash('Пожалуйста, войдите в систему')
        return redirect(url_for('login'))

    required_role = ROLE_REQUIRED.get(endpoint)
    if required_role is not None:
        user_role = session.get('role')
        if user_role != required_role:
            flash('У вас нет прав для доступа к этой странице')
            home = ROLE_HOME.get(user_role, 'login')
            return redirect(url_for(home))

@app.context_processor
def inject_unread_count():
    if 'user_id' in session:
        unread = get_unread_notifications(session['user_id'])
        return {'unread_count': len(unread)}
    return {'unread_count': 0}
@app.route('/')
def index():
    if 'user_id' in session:
        user = get_user_by_id(session['user_id'])
        if user:
            if user.role == 'student':
                return redirect(url_for('student'))
            elif user.role == 'cook':
                return redirect(url_for('cook'))
            elif user.role == 'admin':
                return redirect(url_for('admin'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = get_user(username)
        if user and user.check_password(password):
            if not user.is_approved:
                flash('Ваш аккаунт ещё не подтверждён администратором')
                return render_template('login.html')
            session['user_id'] = user.id
            session['role'] = user.role
            return redirect(url_for('index'))
        flash('Неверный логин или пароль')
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        full_name = request.form.get('full_name', '')
        class_name = request.form.get('class_name', '')

        if get_user(username):
            flash('Пользователь уже существует')
            return render_template('register.html')
        else:
            add_user(username, password, 'student', full_name, class_name)
            admins = User.query.filter_by(role='admin').all()
            for adm in admins:
                add_notification(adm.id, f'Новая заявка на регистрацию: {full_name} ({username})')
            flash('Заявка отправлена. Ожидайте подтверждения.')
            return redirect(url_for('login'))
    return render_template('register.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('Вы вышли из системы')
    return redirect(url_for('login'))


@app.route('/student')
def student():
    user = get_user_by_id(session['user_id'])
    breakfast_menu = get_all_unique_menu_items('breakfast')
    lunch_menu = get_all_unique_menu_items('lunch')
    user_allergies = get_user_allergy_ids(user.id)
    breakfast_sub = get_subscription(user.id, 'breakfast')
    lunch_sub = get_subscription(user.id, 'lunch')
    orders = get_user_orders(user.id)
    unread = get_unread_notifications(user.id)
    unavailable_ids = get_unavailable_item_ids()

    menu_allergies = {}
    all_items = []
    for items in breakfast_menu.values():
        all_items.extend(items)
    for items in lunch_menu.values():
        all_items.extend(items)

    for item in all_items:
        menu_allergies[item.id] = get_menu_item_allergies(item.id)

    sub_prices = {'breakfast': 100, 'lunch': 150}

    return render_template('student.html',
                           user=user,
                           breakfast_menu=breakfast_menu,
                           lunch_menu=lunch_menu,
                           user_allergies=user_allergies,
                           menu_allergies=menu_allergies,
                           breakfast_sub=breakfast_sub,
                           lunch_sub=lunch_sub,
                           orders=orders,
                           unread_count=len(unread),
                           sub_prices=sub_prices,
                           unavailable_ids=unavailable_ids)


@app.route('/order', methods=['POST'])
def order():
    user = get_user_by_id(session['user_id'])
    meal_type = request.form.get('meal_type')
    use_sub = request.form.get('use_subscription') == '1'

    # Собираем отмеченные чекбоксы item_XX
    item_ids = []
    for key in request.form:
        if key.startswith('item_'):
            val = request.form.get(key)
            if val:
                item_ids.append(int(val))

    if not item_ids:
        flash('Выберите блюда')
        return redirect(url_for('student'))

    unavailable = get_unavailable_item_ids()
    for item_id in item_ids:
        if item_id in unavailable:
            flash('Одно из выбранных блюд недоступно')
            return redirect(url_for('student'))

    prices = {'breakfast': 100, 'lunch': 150}
    sub_price = prices.get(meal_type, 100)

    if use_sub:
        today = date.today()
        already_sub_orders = get_subscription_orders_count_for_day(user.id, meal_type, today)
        if already_sub_orders >= 1:
            flash('Вы уже использовали абонемент на этот прием пищи сегодня')
            return redirect(url_for('student'))

        sub = get_subscription(user.id, meal_type)
        if not sub or sub.meals_left < 1:
            flash('Нет доступных посещений в абонементе')
            return redirect(url_for('student'))

        use_subscription(user.id, meal_type)
        order_obj, total = create_order(user.id, meal_type, item_ids, is_subscription=True)
        add_notification(user.id, f'Заказ по абонементу оформлен ({sub_price} руб).')
    else:
        items = [MenuItem.query.get(i) for i in item_ids]
        total = sum(i.price for i in items if i)

        if user.balance < total:
            flash('Недостаточно средств')
            return redirect(url_for('student'))

        subtract_balance(user.id, total)
        add_payment(user.id, total, 'purchase')
        order_obj, total = create_order(user.id, meal_type, item_ids, is_subscription=False)
        add_notification(user.id, f'Заказ на {total} руб. оформлен!')

    cooks = User.query.filter_by(role='cook').all()
    for c in cooks:
        add_notification(c.id, f'Новый заказ #{order_obj.id}')

    flash('Заказ успешно создан!')
    return redirect(url_for('student'))


@app.route('/buy_subscription', methods=['POST'])
def buy_subscription():
    user = get_user_by_id(session['user_id'])
    meal_type = request.form.get('meal_type')
    count = int(request.form.get('count', 5))

    prices = {'breakfast': 100, 'lunch': 150}
    price = prices.get(meal_type, 100) * count

    if user.balance < price:
        flash('Недостаточно средств')
        return redirect(url_for('student'))

    subtract_balance(user.id, price)
    add_payment(user.id, price, 'subscription')
    add_subscription(user.id, meal_type, count)

    meal_name_ru = 'завтраков' if meal_type == 'breakfast' else 'обедов'
    add_notification(user.id, f'Куплен абонемент на {count} {meal_name_ru}!')

    flash(f'Абонемент куплен!')
    return redirect(url_for('student'))


@app.route('/add_balance', methods=['POST'])
def add_balance_route():
    amount = float(request.form.get('amount', 0))
    card_num = request.form.get('card_num', '').replace(' ', '')
    cvc = request.form.get('cvc', '')

    if len(card_num) < 16 or len(cvc) < 3:
        flash('Ошибка оплаты: Неверные данные карты')
        return redirect(url_for('student'))

    if amount > 0:
        add_balance(session['user_id'], amount)
        add_notification(session['user_id'], f'Баланс пополнен на {amount} руб.')
        flash(f'Баланс пополнен!')
    return redirect(url_for('student'))


@app.route('/receive_order/<int:order_id>')
def receive_order(order_id):
    if mark_order_received(order_id, session['user_id']):
        flash('Заказ получен!')
    else:
        flash('Ошибка получения заказа')
    return redirect(url_for('student'))


@app.route('/profile')
def profile():
    user = get_user_by_id(session['user_id'])
    allergies = get_all_allergies()
    user_allergies = get_user_allergy_ids(user.id)
    return render_template('profile.html', user=user, allergies=allergies, user_allergies=user_allergies)


@app.route('/toggle_allergy/<int:allergy_id>')
def toggle_allergy(allergy_id):
    user_allergies = get_user_allergy_ids(session['user_id'])
    if allergy_id in user_allergies:
        remove_user_allergy(session['user_id'], allergy_id)
    else:
        add_user_allergy(session['user_id'], allergy_id)
    return redirect(url_for('profile'))


@app.route('/notifications')
def notifications():
    notifs = get_notifications(session['user_id'])
    mark_all_notifications_read(session['user_id'])
    user = get_user_by_id(session['user_id'])
    return render_template('notifications.html', notifications=notifs, user=user)


@app.route('/reviews')
def reviews():
    user = get_user_by_id(session['user_id'])
    all_reviews = get_all_reviews()
    items = MenuItem.query.all()
    return render_template('reviews.html', user=user, reviews=all_reviews, items=items)


@app.route('/add_review', methods=['POST'])
def add_review_route():
    menu_item_id = int(request.form.get('menu_item_id'))
    review_text = request.form.get('text')
    rating = int(request.form.get('rating', 5))

    add_review(session['user_id'], menu_item_id, review_text, rating)

    item = MenuItem.query.get(menu_item_id)
    user = get_user_by_id(session['user_id'])
    item_name = item.name if item else 'Неизвестное блюдо'
    user_name = user.full_name or user.username
    stars = '★' * rating + '☆' * (5 - rating)

    cooks = User.query.filter_by(role='cook').all()
    for c in cooks:
        add_notification(
            c.id,
            f'Новый отзыв на "{item_name}" от {user_name}: {stars} — "{review_text}"'
        )

    flash('Отзыв добавлен!')
    return redirect(url_for('reviews'))


@app.route('/toggle_favorite/<int:item_id>')
def toggle_favorite(item_id):
    fav = Favorite.query.filter_by(user_id=session['user_id'], menu_item_id=item_id).first()
    if fav:
        db.session.delete(fav)
        flash('Убрано из избранного')
    else:
        new_fav = Favorite(user_id=session['user_id'], menu_item_id=item_id)
        db.session.add(new_fav)
        flash('Добавлено в избранное')
    db.session.commit()
    return redirect(request.referrer or url_for('student'))


@app.route('/cook')
def cook():
    user = get_user_by_id(session['user_id'])
    orders = get_orders_to_prepare()
    products = get_all_products()
    my_requests = PurchaseRequest.query.filter_by(created_by=user.id).order_by(PurchaseRequest.date.desc()).all()
    unread = get_unread_notifications(user.id)

    order_item_can_cook = {}
    for o in orders:
        for oi in o.items:
            if not oi.is_cooked:
                order_item_can_cook[oi.id] = check_item_ingredients_available(oi.menu_item_id)

    today = date.today()
    today_orders = Order.query.filter_by(date=today).all()

    b_count = len([o for o in today_orders if o.meal_type == 'breakfast'])
    l_count = len([o for o in today_orders if o.meal_type == 'lunch'])
    p_count = len([o for o in today_orders if o.is_prepared])
    r_count = len([o for o in today_orders if o.is_received])

    return render_template('cook.html',
                           user=user,
                           orders=orders,
                           products=products,
                           my_requests=my_requests,
                           order_item_can_cook=order_item_can_cook,
                           unread_count=len(unread),
                           breakfast_count=b_count,
                           lunch_count=l_count,
                           prepared_count=p_count,
                           received_count=r_count)


@app.route('/cook/dishes')
def cook_dishes():
    user = get_user_by_id(session['user_id'])
    unread = get_unread_notifications(user.id)
    breakfast_dishes = get_all_unique_menu_items('breakfast')
    lunch_dishes = get_all_unique_menu_items('lunch')

    dish_ingredients = {}
    dish_can_cook = {}
    all_dishes = []
    for items in breakfast_dishes.values():
        all_dishes.extend(items)
    for items in lunch_dishes.values():
        all_dishes.extend(items)

    for item in all_dishes:
        ings = MenuItemIngredient.query.filter_by(menu_item_id=item.id).all()
        dish_ingredients[item.id] = ings
        dish_can_cook[item.id] = item.is_available and check_item_ingredients_available(item.id)

    categories = Category.query.all()
    products = get_all_products()
    allergies = get_all_allergies()

    return render_template('cook_dishes.html',
                           user=user,
                           breakfast_dishes=breakfast_dishes,
                           lunch_dishes=lunch_dishes,
                           dish_ingredients=dish_ingredients,
                           dish_can_cook=dish_can_cook,
                           categories=categories,
                           products=products,
                           allergies=allergies,
                           unread_count=len(unread))


@app.route('/cook/toggle_item/<int:item_id>')
def toggle_item(item_id):
    toggle_menu_item_availability(item_id)
    return redirect(url_for('cook_dishes'))


@app.route('/cook/create_dish', methods=['POST'])
def create_dish():
    name = request.form.get('dish_name', '').strip()
    price = float(request.form.get('dish_price', 0))
    category_id = int(request.form.get('dish_category'))

    ing_ids = request.form.getlist('ing_product[]')
    ing_qtys = request.form.getlist('ing_qty[]')

    ingredients_data = []
    for pid, qty in zip(ing_ids, ing_qtys):
        if pid and qty:
            ingredients_data.append((int(pid), float(qty)))

    allergy_ids = [int(a) for a in request.form.getlist('dish_allergies') if a]

    create_custom_dish(name, price, category_id, ingredients_data, allergy_ids)
    flash('Блюдо создано')
    return redirect(url_for('cook_dishes'))


@app.route('/cook/delete_dish/<int:item_id>')
def delete_dish(item_id):
    delete_menu_item(item_id)
    flash('Блюдо удалено')
    return redirect(url_for('cook_dishes'))


@app.route('/api/cook_item/<int:oi_id>')
def cook_item(oi_id):
    if mark_order_item_cooked(oi_id):
        flash('Блюдо приготовлено')
    else:
        flash('Ошибка или нехватка продуктов')
    return redirect(url_for('cook'))


@app.route('/api/prepare_order/<int:order_id>')
def prepare_order(order_id):
    if mark_order_prepared(order_id):
        order = Order.query.get(order_id)
        add_notification(order.user_id, f'Заказ #{order.id} готов к выдаче!')
        flash('Заказ готов к выдаче')
    else:
        flash('Не все блюда готовы')
    return redirect(url_for('cook'))


@app.route('/add_request', methods=['POST'])
def add_request():
    product_id = int(request.form.get('product_id'))
    quantity = float(request.form.get('quantity'))
    add_purchase_request(product_id, quantity, session['user_id'])

    # Получаем данные для уведомления
    product = Product.query.get(product_id)
    cook_user = get_user_by_id(session['user_id'])
    product_name = product.name if product else 'Неизвестный продукт'
    cook_name = cook_user.full_name or cook_user.username
    unit = product.unit if product else 'ед.'

    # Отправляем уведомление всем админам
    admins = User.query.filter_by(role='admin').all()
    for adm in admins:
        add_notification(
            adm.id,
            f'Новая заявка на закупку от {cook_name}: {product_name} — {quantity} {unit}'
        )

    flash('Заявка создана')
    return redirect(url_for('cook'))

@app.route('/cook/issued')
def cook_issued():
    user = get_user_by_id(session['user_id'])
    unread = get_unread_notifications(user.id)

    today = date.today()
    issued_today = Order.query.filter_by(is_received=True, date=today).all()

    b_issued = [o for o in issued_today if o.meal_type == 'breakfast']
    l_issued = [o for o in issued_today if o.meal_type == 'lunch']

    dish_stats = {}
    for order in issued_today:
        for item in order.items:
            name = item.menu_item.name
            if name not in dish_stats:
                dish_stats[name] = 0
            dish_stats[name] += 1

    dish_stats = dict(sorted(dish_stats.items(), key=lambda item: item[1], reverse=True))

    return render_template('cook_issued.html',
                           user=user,
                           breakfast_issued=b_issued,
                           lunch_issued=l_issued,
                           dish_stats=dish_stats,
                           unread_count=len(unread))


@app.route('/admin')
def admin():
    user = get_user_by_id(session['user_id'])
    stats = get_payments_stats()
    order_stats = get_orders_stats()
    expenses = get_expenses()
    pending = get_pending_requests()
    users = User.query.all()
    pending_users = User.query.filter_by(is_approved=False).all()
    unread = get_unread_notifications(user.id)

    today = date.today()
    all_today = Order.query.filter_by(date=today).all()

    class_stats = {}
    for o in all_today:
        if o.is_received:
            u_class = o.user.class_name or 'Без класса'
            if u_class not in class_stats:
                class_stats[u_class] = {'total': 0, 'breakfast': 0, 'lunch': 0}
            class_stats[u_class]['total'] += 1
            if o.meal_type == 'breakfast':
                class_stats[u_class]['breakfast'] += 1
            else:
                class_stats[u_class]['lunch'] += 1

    total_b = len([o for o in all_today if o.meal_type == 'breakfast'])
    total_l = len([o for o in all_today if o.meal_type == 'lunch'])

    return render_template('admin.html',
                           user=user,
                           stats=stats,
                           order_stats=order_stats,
                           expenses=expenses,
                           pending=pending,
                           users=users,
                           pending_users=pending_users,
                           class_stats=class_stats,
                           unread_count=len(unread),
                           total_breakfasts=order_stats['total'],
                           breakfast_today=total_b,
                           total_lunches=order_stats['total'],
                           lunch_today=total_l)


@app.route('/approve/<int:req_id>')
def approve(req_id):
    if approve_request(req_id):
        flash('Заявка одобрена')
    return redirect(url_for('admin'))


@app.route('/reject/<int:req_id>')
def reject(req_id):
    if reject_request(req_id):
        flash('Заявка отклонена')
    return redirect(url_for('admin'))


@app.route('/approve_user/<int:user_id>')
def approve_user(user_id):
    u = User.query.get(user_id)
    if u:
        u.is_approved = True
        db.session.commit()
        add_notification(u.id, 'Аккаунт подтвержден')
        flash('Пользователь подтвержден')
    return redirect(url_for('admin'))


@app.route('/reject_user/<int:user_id>')
def reject_user(user_id):
    u = User.query.get(user_id)
    if u and not u.is_approved:
        db.session.delete(u)
        db.session.commit()
        flash('Заявка отклонена')
    return redirect(url_for('admin'))


@app.route('/download_report')
def download_report():
    stats = get_payments_stats()
    order_stats = get_orders_stats()
    expenses = get_expenses()
    all_requests = get_all_requests()
    users = User.query.order_by(User.id).all()
    all_orders = Order.query.order_by(Order.created_at.desc()).all()
    products = Product.query.order_by(Product.id).all()
    all_order_items = OrderItem.query.all()
    menu_items = MenuItem.query.order_by(MenuItem.name).all()

    wb = Workbook()

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Side(style='thin')
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    center_align = Alignment(horizontal='center', vertical='center')
    title_font = Font(name='Calibri', size=14, bold=True)
    bold_font = Font(name='Calibri', bold=True)

    def style_table_header(ws, row, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

    ws1 = wb.active
    ws1.title = "Сводка"
    ws1['A1'] = "Отчёт по школьной столовой"
    ws1['A1'].font = title_font
    ws1['A2'] = "Дата формирования"
    ws1['B2'] = datetime.now().strftime('%d.%m.%Y %H:%M')

    ws1['A4'] = "Финансовые показатели"
    ws1['A4'].font = bold_font

    ws1.append(['Показатель', 'Сумма (руб.)'])
    style_table_header(ws1, 5, 2)

    data_fin = [
        ['Доход от абонементов', stats['subscriptions']],
        ['Доход от покупок', stats['purchases']],
        ['Общий доход', stats['total_income']],
        ['Расходы на закупки', expenses],
        ['Чистая прибыль', stats['total_income'] - expenses]
    ]
    for row in data_fin:
        ws1.append(row)

    ws1['A12'] = "Статистика заказов"
    ws1['A12'].font = bold_font

    ws1.append(['Показатель', 'Значение'])
    style_table_header(ws1, 13, 2)

    b_total = len([o for o in all_orders if o.meal_type == 'breakfast'])
    l_total = len([o for o in all_orders if o.meal_type == 'lunch'])

    data_ord = [
        ['Всего заказов', order_stats['total']],
        ['Заказов за сегодня', order_stats['today']],
        ['Получено', order_stats['received']],
        ['Всего завтраков', b_total],
        ['Всего обедов', l_total]
    ]
    for row in data_ord:
        ws1.append(row)

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet("Учёт завтраков")
    ws2['A1'] = "Учёт завтраков"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Всего завтраков:"
    ws2['B2'] = b_total

    headers_meals = ['№', 'Дата', 'Ученик', 'Класс', 'Блюда', 'Оплата', 'Статус']
    ws2.append([])
    ws2.append(headers_meals)
    style_table_header(ws2, 4, 7)

    for o in all_orders:
        if o.meal_type == 'breakfast':
            items_str = ", ".join([i.menu_item.name for i in o.items])
            pay_type = 'Абонемент' if o.is_subscription else 'Разовая'
            status = 'Получен' if o.is_received else ('Готов' if o.is_prepared else 'Готовится')
            ws2.append([o.id, o.date.strftime('%d.%m.%Y'), o.user.full_name or o.user.username,
                        o.user.class_name, items_str, pay_type, status])

    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['E'].width = 40

    ws3 = wb.create_sheet("Учёт обедов")
    ws3['A1'] = "Учёт обедов"
    ws3['A1'].font = title_font
    ws3['A2'] = f"Всего обедов:"
    ws3['B2'] = l_total

    ws3.append([])
    ws3.append(headers_meals)
    style_table_header(ws3, 4, 7)

    for o in all_orders:
        if o.meal_type == 'lunch':
            items_str = ", ".join([i.menu_item.name for i in o.items])
            pay_type = 'Абонемент' if o.is_subscription else 'Разовая'
            status = 'Получен' if o.is_received else ('Готов' if o.is_prepared else 'Готовится')
            ws3.append([o.id, o.date.strftime('%d.%m.%Y'), o.user.full_name or o.user.username,
                        o.user.class_name, items_str, pay_type, status])

    ws3.column_dimensions['C'].width = 25
    ws3.column_dimensions['E'].width = 40

    ws4 = wb.create_sheet("Все заказы")
    ws4['A1'] = "Все заказы"
    ws4['A1'].font = title_font

    headers_all = ['№', 'Дата', 'Ученик', 'Класс', 'Приём пищи', 'Блюда', 'Оплата', 'Статус']
    ws4.append([])
    ws4.append(headers_all)
    style_table_header(ws4, 3, 8)

    for o in all_orders:
        meal = 'Завтрак' if o.meal_type == 'breakfast' else 'Обед'
        items_str = ", ".join([i.menu_item.name for i in o.items])
        pay_type = 'Абонемент' if o.is_subscription else 'Разовая'
        status = 'Получен' if o.is_received else ('Готов' if o.is_prepared else 'Готовится')
        ws4.append([o.id, o.date.strftime('%d.%m.%Y'), o.user.full_name or o.user.username,
                    o.user.class_name, meal, items_str, pay_type, status])

    ws4.column_dimensions['C'].width = 25
    ws4.column_dimensions['F'].width = 40

    ws5 = wb.create_sheet("Заявки на закупку")
    ws5['A1'] = "Заявки на закупку"
    ws5['A1'].font = title_font

    headers_req = ['Дата', 'Продукт', 'Количество', 'Ед.', 'Сумма (руб.)']
    ws5.append([])
    ws5.append(headers_req)
    style_table_header(ws5, 3, 5)

    for r in all_requests:
        summ = round(r.quantity * r.product.price, 2)
        ws5.append([r.date.strftime('%d.%m.%Y'), r.product.name, r.quantity, r.product.unit, summ])

    ws5.column_dimensions['B'].width = 25

    ws6 = wb.create_sheet("Пользователи")
    ws6['A1'] = "Пользователи"
    ws6['A1'].font = title_font

    headers_users = ['ID', 'Логин', 'ФИО', 'Класс', 'Роль']
    ws6.append([])
    ws6.append(headers_users)
    style_table_header(ws6, 3, 5)

    for u in users:
        role_map = {'student': 'Ученик', 'cook': 'Повар', 'admin': 'Администратор'}
        ws6.append([u.id, u.username, u.full_name, u.class_name, role_map.get(u.role, u.role)])

    ws6.column_dimensions['C'].width = 30

    ws7 = wb.create_sheet("Продукты")
    ws7['A1'] = "Продукты"
    ws7['A1'].font = title_font

    headers_prod = ['ID продукта', 'Название', 'Остаток', 'Ед. измерения']
    ws7.append([])
    ws7.append(headers_prod)
    style_table_header(ws7, 3, 4)

    for p in products:
        ws7.append([p.id, p.name, p.quantity, p.unit])

    ws7.column_dimensions['B'].width = 25

    ws8 = wb.create_sheet("Приготовленные блюда")
    ws8['A1'] = "Приготовленные блюда"
    ws8['A1'].font = title_font

    headers_cooked = ['ID позиции', 'ID заказа', 'ID блюда', 'Блюдо']
    ws8.append([])
    ws8.append(headers_cooked)
    style_table_header(ws8, 3, 4)

    for oi in all_order_items:
        ws8.append([oi.id, oi.order_id, oi.menu_item_id, oi.menu_item.name])

    ws8.column_dimensions['D'].width = 30

    ws9 = wb.create_sheet("Связь блюд и продуктов")
    ws9['A1'] = "Связь блюд и продуктов"
    ws9['A1'].font = title_font

    headers_link = ['ID блюда', 'Блюдо', 'ID продукта', 'Продукт']
    ws9.append([])
    ws9.append(headers_link)
    style_table_header(ws9, 3, 4)

    for item in menu_items:
        ings = MenuItemIngredient.query.filter_by(menu_item_id=item.id).all()
        for ing in ings:
            ws9.append([item.id, item.name, ing.product_id, ing.product.name])

    ws9.column_dimensions['B'].width = 25
    ws9.column_dimensions['D'].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='school_canteen_report.xlsx', as_attachment=True)


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0')